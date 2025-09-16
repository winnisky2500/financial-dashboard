# simulation_agent.py
import os, io, csv, json, hashlib, datetime as dt, re
from typing import List, Dict, Any, Optional
from uuid import uuid4

from fastapi import FastAPI, UploadFile, File, Form, Body, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

from supabase import create_client, Client
import numpy as np

# ARIMA
from statsmodels.tsa.arima.model import ARIMA

# 新增：导出 XLSX（带公式与加粗分隔）
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# =============== 环境与客户端 ===============
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
if not SUPABASE_URL or not SUPABASE_KEY:
    raise RuntimeError("Missing SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY")

sb: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# （可选）OpenAI，用于 LLM 提示
import openai
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
if OPENAI_API_KEY:
    openai.api_key = OPENAI_API_KEY

GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY")
GOOGLE_CSE_ID = os.getenv("GOOGLE_CSE_ID")
app = FastAPI(title="Simulation Agent V2")

# 仅允许本地前端来源；带凭据时不能使用 "*"
ALLOWED_ORIGINS = [
    "http://localhost:5173",
    "http://127.0.0.1:5173",
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# =============== 数据模型 ===============
class SeedRequest(BaseModel):
    question: str
    run_id: Optional[str] = None
    # 由前端传入：通常用 supabase 的 session.user.id
    session_user_id: Optional[str] = None

class SensitivityRow(BaseModel):
    company_name: str
    canonical_metric: str
    factor_name: str
    elasticity_value: Optional[float] = None
    lag_quarters: int = 0
    shock_unit: str = "percent"
    source_method: Optional[str] = "user"
    note: Optional[str] = None
    # 季节性（按 Y 指标）
    seasonal_adjust: Optional[bool] = None
    seasonality_source: Optional[str] = None
    seasonality_q1: Optional[float] = None
    seasonality_q2: Optional[float] = None
    seasonality_q3: Optional[float] = None
    seasonality_q4: Optional[float] = None



class ScenarioDelta(BaseModel):
    factor: str
    optimistic: float = 0.05  # 乐观：+5%
    base: float = 0.0         # 平缓/基准：0%
    pessimistic: float = -0.05# 悲观：-5%

class ModelsConfig(BaseModel):
    arima: Dict[str, Any] = {"enabled": True, "p": 1, "d": 1, "q": 1, "periods": 8}
    monte_carlo: Dict[str, Any] = {"enabled": True, "samples": 1000, "quantiles": [0.1, 0.5, 0.9]}

class RunRequest(BaseModel):
    run_id: str
    sensitivity_rows: List[SensitivityRow]
    models: ModelsConfig
    horizon_quarters: int = 8
    # 由前端传入：通常用 supabase 的 session.user.id
    session_user_id: Optional[str] = None
    # 快速再生成：仅数学推演，不重新写 MD 报告
    skip_report: Optional[bool] = False
    # 新增：情景自变量冲击（用户/LLM可编辑）
    scenario_deltas: Optional[List[ScenarioDelta]] = None

class SaveMdRequest(BaseModel):
    run_id: str
    md: str
    session_user_id: Optional[str] = None
@app.post("/simulation_v2/save_md")
def save_md(req: SaveMdRequest):
    # run 存在性兜底（可选）
    upsert_run(
        req.run_id,
        title=f"模拟分析2 - 手工更新报告",
        models={"note": "manual-md-update"},
        session_id=req.session_user_id or "manual"
    )
    md_bytes = req.md.encode("utf-8")
    url_md = save_artifact(req.run_id, "md", md_bytes, "report.md")
    return {"ok": True, "report_url": url_md}

# === 新增：把 MD 交给 beautifyreport_agent 渲染（HTML/PDF），再回存到 Storage ===
@app.post("/simulation_v2/beautify_md")
def beautify_md(req: SaveMdRequest):
    import httpx
    md_text = req.md
    # 外部美化服务地址（没有就直接回存原文）
    target = os.getenv("BEAUTIFY_AGENT_URL")  # 例如 http://127.0.0.1:18071
    if not target:
        url_md = save_artifact(req.run_id, "md", md_text.encode("utf-8"), "report.md")
        return {"ok": True, "report_url": url_md, "notice": "BEAUTIFY_AGENT_URL 未配置，已仅保存MD"}

    # 调用美化服务
    with httpx.Client(timeout=60) as cli:
        r = cli.post(f"{target}/render_md", json={"run_id": req.run_id, "md": md_text})
        r.raise_for_status()
        data = r.json()

    # 期望 beautify 返回 html 或 pdf 原文（这里兼容两种）
    if "html" in data and isinstance(data["html"], str):
        html_bytes = data["html"].encode("utf-8")
        url_html = save_artifact(req.run_id, "html", html_bytes, "report.html")
        return {"ok": True, "report_url": url_html, "type": "html"}
    if "pdf_base64" in data and isinstance(data["pdf_base64"], str):
        import base64
        pdf_bytes = base64.b64decode(data["pdf_base64"])
        url_pdf = save_artifact(req.run_id, "pdf", pdf_bytes, "report.pdf")
        return {"ok": True, "report_url": url_pdf, "type": "pdf"}

    # 回退：仅保存 MD
    url_md = save_artifact(req.run_id, "md", md_text.encode("utf-8"), "report.md")
    return {"ok": True, "report_url": url_md, "type": "md-fallback"}
# =============== 工具函数 ===============
def sha256_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()

def now_iso() -> str:
    return dt.datetime.now(dt.timezone.utc).isoformat()

def arima_baseline(series: List[float], p: int, d: int, q: int, periods: int) -> Dict[str, Any]:
    # 简易 ARIMA；失败则平滑外推
    try:
        model = ARIMA(series, order=(p, d, q))
        res = model.fit()
        forecast = list(res.forecast(steps=periods))
        conf_int = res.get_forecast(steps=periods).conf_int(alpha=0.05).tolist()
        lower95 = [float(x[0]) for x in conf_int]
        upper95 = [float(x[1]) for x in conf_int]
        return {"pred": forecast, "lower95": lower95, "upper95": upper95}
    except Exception:
        if not series:
            return {"pred": [0.0]*periods, "lower95": [0.0]*periods, "upper95": [0.0]*periods}
        last = series[-1]
        return {"pred": [last for _ in range(periods)], "lower95": [last for _ in range(periods)], "upper95": [last for _ in range(periods)]}
def google_search_snippets(query: str, topk: int = 3) -> List[Dict[str, str]]:
    if not GOOGLE_API_KEY or not GOOGLE_CSE_ID:
        return []
    try:
        import httpx
        params = {"key": GOOGLE_API_KEY, "cx": GOOGLE_CSE_ID, "q": query}
        with httpx.Client(timeout=20) as cli:
            r = cli.get("https://www.googleapis.com/customsearch/v1", params=params)
            r.raise_for_status()
            data = r.json()
        items = data.get("items") or []
        out = []
        for it in items[:topk]:
            out.append({
                "title": it.get("title",""),
                "snippet": it.get("snippet",""),
                "link": it.get("link","")
            })
        return out
    except Exception:
        return []
def load_series_rows(company: str, metric: str, max_points: int = 40) -> List[Dict[str, Any]]:
    rows = (sb.table("financial_metrics")
            .select("company_name, metric_name, year, quarter, metric_value")
            .eq("company_name", company)
            .eq("metric_name", metric)
            .order("year", desc=False)
            .order("quarter", desc=False)
            .limit(max_points)
            .execute().data or [])
    return rows

def q_back(y: int, q: int, steps: int) -> List[tuple]:
    out = []
    yy, qq = y, q
    for _ in range(steps):
        out.append((yy, qq))
        qq -= 1
        if qq == 0:
            yy -= 1; qq = 4
    return list(reversed(out))

def q_forward(y: int, q: int, steps: int, mark_est: bool = True) -> List[str]:
    out = []
    yy, qq = y, q
    for _ in range(steps):
        # 下一期
        qq += 1
        if qq == 5:
            yy += 1; qq = 1
        lab = f"{yy}Q{qq}"
        if mark_est:
            lab += "(e)"
        out.append(lab)
    return out
def llm_infer_sensitivity(company: str, y_metric: str, x_metric: str, run_id: Optional[str] = None) -> Dict[str, Any]:
    """
    数据→(QoQ/YoY)→最小二乘 OLS 给出“统计参考β”，然后把该统计结果 + 附件摘要交给 LLM，请其在统计值附近做口径校准。
    返回：{"beta": float or None, "rationale": str, "method": "qoq|yoy|expert"}
    """
    # 1) 原始行
    y_rows = load_series_rows(company, y_metric, max_points=24)
    x_rows = load_series_rows(company, x_metric, max_points=24)

    def _to_map(rows):
        return {(int(r["year"]), int(r["quarter"])): float(r["metric_value"]) for r in rows}

    ym = _to_map(y_rows); xm = _to_map(x_rows)

    # 2) 组装 QoQ/YoY 百分比变化序列（对齐相同时间点，且要能取到 t-1 或 t-4）
    def pct(v_now, v_prev):
        try:
            v_now = float(v_now); v_prev = float(v_prev)
            if v_prev == 0: return None
            return (v_now - v_prev) / abs(v_prev)
        except Exception:
            return None

    qoq_X, qoq_Y = [], []
    yoy_X, yoy_Y = [], []
    # 列举所有 (year, quarter) 键
    keys = sorted(set(list(ym.keys()) + list(xm.keys())))
    for (yy, qq) in keys:
        # QoQ
        prev = (yy, qq-1) if qq>1 else (yy-1, 4)
        if (yy, qq) in ym and (yy, qq) in xm and prev in ym and prev in xm:
            dy = pct(ym[(yy,qq)], ym[prev]); dx = pct(xm[(yy,qq)], xm[prev])
            if dy is not None and dx is not None:
                qoq_X.append(dx); qoq_Y.append(dy)
        # YoY
        prev4 = (yy-1, qq)
        if (yy, qq) in ym and (yy, qq) in xm and prev4 in ym and prev4 in xm:
            dy = pct(ym[(yy,qq)], ym[prev4]); dx = pct(xm[(yy,qq)], xm[prev4])
            if dy is not None and dx is not None:
                yoy_X.append(dx); yoy_Y.append(dy)

    def ols_beta(xs, ys):
        if not xs or not ys or len(xs) != len(ys):
            return None, None
        X = np.array(xs, dtype=float); Y = np.array(ys, dtype=float)
        # 通过原点拟合：Y ≈ β*X
        try:
            beta = float(np.linalg.lstsq(X.reshape(-1,1), Y, rcond=None)[0][0])
            corr = float(np.corrcoef(X, Y)[0,1])
            return beta, corr
        except Exception:
            return None, None

    beta_qoq, corr_qoq = ols_beta(qoq_X, qoq_Y)
    beta_yoy, corr_yoy = ols_beta(yoy_X, yoy_Y)

    # 3) 附件（CSV）摘要（与原逻辑相同）
    attach_snips = []
    if run_id:
        try:
            atts = sb.table("run_attachments").select("*").eq("run_id", run_id).order("id", desc=True).limit(5).execute().data or []
            for a in atts:
                url = a.get("storage_url"); fname = a.get("filename")
                if url and fname and fname.lower().endswith(".csv"):
                    try:
                        import httpx
                        with httpx.Client(timeout=15) as cli:
                            t = cli.get(url).text.splitlines()[:5]
                        attach_snips.append({"filename": fname, "preview": "\n".join(t)})
                    except Exception:
                        attach_snips.append({"filename": fname, "preview": "(读取失败，忽略)"})
        except Exception:
            pass

    # 4) 无 Key：直接返回 OLS 结果（或 None）
    if not OPENAI_API_KEY:
        # 优先 YoY，再 QoQ
        beta = beta_yoy if beta_yoy is not None else beta_qoq
        return {"beta": beta, "rationale": "LLM disabled. Returned OLS estimate.", "method": "yoy" if beta==beta_yoy else "qoq"}

    # 5) 让 LLM 在统计值基础上校准（优先改提示词）
    sys_prompt = """你是财务灵敏度估计助手。
现在给你 X 与 Y 的 QoQ/YoY 百分比变化序列的最小二乘回归结果（通过原点）：beta_qoq/beta_yoy 及对应相关系数。
请结合这些统计结果 + 附件片段，产出一个“合理且可解释”的 β。若两者差异大，请给出选择理由。
只返回JSON：
{"beta": 0.12, "method":"yoy|qoq|expert", "rationale": "…"}"""

    payload = {
        "company": company,
        "Y_metric": y_metric, "X_metric": x_metric,
        "stats": {
            "beta_qoq": beta_qoq, "corr_qoq": corr_qoq,
            "beta_yoy": beta_yoy, "corr_yoy": corr_yoy,
            "n_qoq": len(qoq_X), "n_yoy": len(yoy_X)
        },
        "attachments_preview": attach_snips
    }

    try:
        txt = call_llm(sys_prompt, json.dumps(payload, ensure_ascii=False))
        data = json.loads(txt)
        beta = data.get("beta")
        method = data.get("method") or ("yoy" if beta_yoy is not None else "qoq")
        beta = float(beta) if beta is not None else (beta_yoy if beta_yoy is not None else beta_qoq)
        return {"beta": beta, "rationale": data.get("rationale",""), "method": method}
    except Exception:
        # 失败时回退到 OLS
        beta = beta_yoy if beta_yoy is not None else beta_qoq
        return {"beta": beta, "rationale": "LLM parse fail. Fallback to OLS.", "method": "yoy" if beta==beta_yoy else "qoq"}


def monte_carlo_paths(baseline: List[float],
                      sensitivity_rows: List[SensitivityRow],
                      samples: int,
                      horizon: int,
                      deltas: Optional[Dict[str, Dict[str, float]]] = None,
                      label: str = "base") -> Dict[str, List[float]]:
    """
    采用“乘法合成”：Y_hat_t * Π_k (1 + beta_k * ΔX_k)。
    这里将 ΔX_k 取自情景字典 deltas[factor][label]，默认 label='base'。
    """
    T = horizon
    base = np.array(baseline, dtype=float)
    mat = np.zeros((samples, T), dtype=float)
    mat += base
    for s in range(samples):
        factor_mult = np.ones(T, dtype=float)
        for row in sensitivity_rows:
            beta = row.elasticity_value or 0.0
            if beta == 0.0:
                continue
            sigma = max(0.0001, 0.2 * abs(beta))
            beta_draw = np.random.normal(beta, sigma)
            lag = max(0, int(row.lag_quarters or 0))

            # 以情景的 base 作为中心，不再硬编码 0.1
            dx = 0.0
            if row.shock_unit == "percent":
                if deltas and row.factor_name in deltas:
                    dx = float(deltas[row.factor_name].get(label, 0.0))

            shock_vec = np.array([(1.0 + beta_draw * dx)] * T)
            if lag > 0:
                shock_vec[:lag] = 1.0
            factor_mult *= shock_vec
        mat[s, :] = base * factor_mult

    pcts = {
        "p10": list(np.percentile(mat, 10, axis=0)),
        "p50": list(np.percentile(mat, 50, axis=0)),
        "p90": list(np.percentile(mat, 90, axis=0)),
    }
    return pcts

def pick_default_Y_from_catalog(catalog: List[str], k: int = 3) -> List[str]:
    """
    从 metric_alias_catalog 的 canonical 列里，按常见财务优先词挑出若干默认 Y。
    优先词包含：净利润/净利率/利润总额/营业收入/毛利率/EBIT。
    """
    pri = ["净利润", "净利率", "利润总额", "营业收入", "毛利率", "EBIT"]
    picked = []
    for p in pri:
        for c in catalog:
            if c not in picked and p in c:
                picked.append(c)
                if len(picked) >= k:
                    return picked
    # 兜底：取前 k 个
    return catalog[:k]

def llm_infer_seasonality(company: str, y_metric: str, run_id: Optional[str] = None) -> Dict[str, Any]:
    """
    让 LLM 阅读 financial_metrics 的季度序列（含 year/quarter），并结合附件摘要判断是否需要季调及强度。
    返回: {"seasonal_adjust": bool,  "rationale": str}
    """
    rows = load_series_rows(company, y_metric, max_points=24)
    series = [{"year": r["year"], "q": r["quarter"], "value": r["metric_value"]} for r in rows]
    attach_snips = []
    if run_id:
        try:
            atts = sb.table("run_attachments").select("*").eq("run_id", run_id).order("id", desc=True).limit(5).execute().data or []
            for a in atts:
                url = a.get("storage_url"); fname = a.get("filename")
                if url and fname and fname.lower().endswith(".csv"):
                    import httpx
                    with httpx.Client(timeout=15) as cli:
                        t = cli.get(url).text.splitlines()[:5]
                    attach_snips.append({"filename": fname, "preview": "\n".join(t)})
        except Exception:
            pass

    if not OPENAI_API_KEY:
        return {"seasonal_adjust": False,  "rationale": "LLM disabled."}

    sys = """你是季节性判断助手。给定季度时间序列(含年/季)和附件片段，
判断是否存在显著季节性。若能估到每季乘数系数，请返回 q1~q4（>1 表示该季偏高，<1 偏低；无法估计就留空）。
仅返回JSON：{"seasonal_adjust": true, "q1":0.95, "q2":1.05, "q3":1.10, "q4":0.90, "rationale":"…"}"""

    user = json.dumps({"company": company, "metric": y_metric, "series": series, "attachments_preview": attach_snips}, ensure_ascii=False)
    try:
        txt = call_llm(sys, user); d = json.loads(txt)
        return {
            "seasonal_adjust": bool(d.get("seasonal_adjust", False)),
            "q1": d.get("q1"), "q2": d.get("q2"), "q3": d.get("q3"), "q4": d.get("q4"),
            "rationale": d.get("rationale","")
            }
    except Exception:
        return {"seasonal_adjust": False, "rationale": ""}

def seasonal_adjust_quarterly_series(company: str, metric: str) -> List[float]:
    """
    简易去季调：对同一季度求均值，得到季节效应 s_q；返回 value - (s_q - overall_mean)。
    """
    rows = load_series_rows(company, metric, max_points=40)
    if not rows: return []
    vals = [float(r["metric_value"]) for r in rows]
    qs = [int(r["quarter"]) for r in rows]
    overall = float(np.mean(vals))
    s = {}
    for q in (1,2,3,4):
        q_vals = [v for v,qq in zip(vals, qs) if qq==q]
        s[q] = float(np.mean(q_vals)) if q_vals else overall
    adj = [v - (s[qq]-overall) for v,qq in zip(vals, qs)]
    return adj
def seasonal_adjust_quarterly_series_with_coeffs(company: str, metric: str,
                                                 q1: Optional[float], q2: Optional[float],
                                                 q3: Optional[float], q4: Optional[float]) -> List[float]:
    """
    使用给定的每季系数进行乘法季节调整：adj = value / q_coef。
    若某季系数缺失，则退回到均值法。
    """
    q1 = 1.0 if q1 is None else float(q1)
    q2 = 1.0 if q2 is None else float(q2)
    q3 = 1.0 if q3 is None else float(q3)
    q4 = 1.0 if q4 is None else float(q4)

    rows = load_series_rows(company, metric, max_points=40)
    if not rows:
        return []
    qs   = [int(r["quarter"]) for r in rows]
    vals = [float(r["metric_value"]) for r in rows]
    coef = {1: q1, 2: q2, 3: q3, 4: q4}
    return [(v / (coef[q] if coef[q] != 0 else 1.0)) for v, q in zip(vals, qs)]

def compute_additive_offsets(company: str, metric: str) -> Dict[int, float]:
    rows = load_series_rows(company, metric, max_points=40)
    if not rows: return {1:0.0,2:0.0,3:0.0,4:0.0}
    vals = [float(r["metric_value"]) for r in rows]
    qs = [int(r["quarter"]) for r in rows]
    overall = float(np.mean(vals))
    offs = {}
    for qq in (1,2,3,4):
        q_vals = [v for v, q in zip(vals, qs) if q == qq]
        offs[qq] = (float(np.mean(q_vals)) - overall) if q_vals else 0.0
    return offs

def reseasonalize_series(base_pred: List[float], start_year: int, start_quarter: int,
                         mode: str, qcoef: Optional[Dict[int,float]] = None,
                         add_offsets: Optional[Dict[int,float]] = None) -> List[float]:
    out = []
    y, q = start_year, start_quarter
    for v in base_pred:
        q += 1
        if q == 5: y += 1; q = 1
        if mode == "multiplicative" and qcoef:
            out.append(float(v) * float(qcoef.get(q, 1.0)))
        elif mode == "additive" and add_offsets:
            out.append(float(v) + float(add_offsets.get(q, 0.0)))
        else:
            out.append(float(v))
    return out

# =============== LLM Prompt 工程 ===============
SYS_PARSE = """你是企业级模拟分析智能体。请从“用户问题+指标字典”中抽取并返回JSON。严格遵循：
- company: 字符串
- X: 字符串数组（若用户只说了一个自变量，也必须以数组返回）；**必须**属于 metric_alias_catalog 中的 canonical_name 列
- Y: 字符串数组；**必须**属于 metric_alias_catalog 中的 canonical_name 列，优先依据 description 列 选择与自变量X最相关的canonical_name；若指标含 is_derived 字段，请**优先选择 is_derived=false** 的候选；并在可行时选择**粒度更细**的指标（例如，若有多个营业收入相关指标，某个地区/子公司/业务/航线的营业收入优先）
- explicit_shocks: 若用户在问题中明确给出幅度/方向（如“关税上调3%/降2bp/价格+5%”），抽取为
  [{"factor":"原文中的自变量称呼","delta":0.03,"unit":"percent","direction":"up","evidence":"原句/理由"}]，否则 []
- notes: 简短说明
- thinking: 一个字符串数组，用第一人称、以“我正在…”开头，描述正在做什么、做到哪一步（每项<=40字，3-8条）
仅返回JSON，不要解释。
"""


SYS_SCENARIO = """你是情景工程师。根据“用户问题/上下文/explicit_shocks/外部检索”，为每个自变量X生成“乐观/平缓/悲观”的ΔX（百分比）与理由。
规则：
- 若 explicit_shocks 中给出了该因子的明确幅度，则 base 必须等于该数值；pessimistic=base*0.5，optimistic=base*1.5（保留符号）。
- 若未给出，则按常识给出小幅度假设（默认 ±5%，基准 0%）。
- 返回：
{"assumptions":[{"factor":"X1","optimistic":0.05,"base":0.03,"pessimistic":0.015,"rationale":"…"}],
 "thinking":["我正在读取显式冲击…","我在对齐因子名称…"]}
仅返回JSON。
"""


SYS_REPORT = """你是报告撰写助手。请输出结构化 Markdown（300-600字），并用清晰小标题说明：
- 问题与假设
- 方法与参数（ARIMA/MonteCarlo）
- 敏感性设定与数据来源（数据库/LLM生成/用户填入）
- 三种情景（乐观/平缓/悲观）关键结论与风险提示
"""


def call_llm(system_prompt: str, user_content: str) -> str:
    if not OPENAI_API_KEY:
        # 无 Key 时返回最小占位
        return '{"company": "", "X": [], "Y": [], "notes": "LLM disabled."}'
    resp = openai.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role":"system","content":system_prompt},
                  {"role":"user","content":user_content}],
        temperature=0.2,
    )
    return resp.choices[0].message.content

# =============== Supabase 访问 ===============
def sb_table_all(table: str) -> List[Dict[str, Any]]:
    return sb.table(table).select("*").execute().data

def find_company_by_name_or_best(question: str) -> str:
    rows = sb.table("company_catalog").select("*").execute().data or []
    text = (question or "").lower()
    best = (rows[0].get("display_name") if rows else None) or "XX集团公司"
    for r in rows:
        dn = (r.get("display_name") or "").strip()
        if not dn: 
            continue
        if dn.lower() in text:
            return dn
        # 兼容 aliases: 数组或以逗号/斜杠等分隔的字符串
        al = r.get("aliases") or r.get("alias") or ""
        cand = al if isinstance(al, list) else re.split(r"[,，/、|;；\s]+", str(al))
        for a in cand:
            if a and str(a).strip().lower() in text:
                return dn
    return best

def as_list(x) -> List[str]:
    """把 None/字符串/标量/数组稳妥转为字符串数组；用于防止 '平均关税税率' 被当成字符序列。"""
    if x is None:
        return []
    if isinstance(x, list):
        return [str(i).strip() for i in x if str(i).strip()]
    if isinstance(x, (dict, tuple, set)):
        return [json.dumps(x, ensure_ascii=False)]
    return [str(x).strip()]

# 在文件顶部的 import 行补上 ", re"（见【改动 2】），此函数依赖正则分割
def alias_to_canonical(names: List[str]) -> List[str]:
    if not names:
        return []
    rows = sb.table("metric_alias_catalog").select("*").execute().data or []

    # 小工具：统一 key
    def _key(x: Any) -> str:
        return str(x).strip().lower()

    # 支持把 "别名1,别名2/别名3；别名4" 这类分隔串拆开
    SPLIT = re.compile(r"[,，/、|;；\s]+")
    def _split_tokens(s: Any) -> List[str]:
        if s is None:
            return []
        parts = [p.strip() for p in SPLIT.split(str(s)) if p and p.strip()]
        return parts or [str(s).strip()]  # 如果没有命中分隔符，就按单值处理

    # 建索引：alias/aliases → canonical_name
    idx: Dict[str, str] = {}
    for r in rows:
        cn = (r.get("canonical_name") or "").strip()
        if not cn:
            continue
        # 自身也允许匹配
        idx[_key(cn)] = cn

        # 单值 alias
        a1 = r.get("alias")
        for tok in _split_tokens(a1):
            idx[_key(tok)] = cn

        # 数组 aliases
        arr = r.get("aliases") or []
        if isinstance(arr, (list, tuple)):
            for tok in arr:
                for t in _split_tokens(tok):
                    idx[_key(t)] = cn

    # 映射输入
    out: List[str] = []
    for nm in names:
        k = _key(nm)
        out.append(idx.get(k, nm))  # 命中则回 canonical，否则保留原词
    # 去重保序
    return list(dict.fromkeys(out))


def list_candidate_Y() -> List[str]:
    rows = sb.table("metric_alias_catalog").select("canonical_name").execute().data
    return sorted(list({r["canonical_name"] for r in rows}))

def load_series_from_financial_metrics(company: str, metric: str, max_points: int = 40) -> List[float]:
    """
    根据公司名 + 指标名取时间序列。
    注意：financial_metrics 表没有 canonical_name，使用 metric_name 对齐。
    这里的 metric 传入值为“canonical 指标名”，需与 financial_metrics.metric_name 一致。
    """
    rows = (sb.table("financial_metrics")
            .select("company_name, metric_name, year, quarter, metric_value")
            .eq("company_name", company)
            .eq("metric_name", metric)      # ← 改用 metric_name
            .order("year", desc=False)
            .order("quarter", desc=False)
            .limit(max_points)
            .execute().data)
    return [float(r["metric_value"]) for r in rows] if rows else []

def lookup_sensitivity(company: str, metric: str, factor: str) -> Optional[float]:
    rows = (sb.table("sensitivity_analysis")
            .select("*")
            .eq("company_name", company)
            .eq("canonical_metric", metric)
            .eq("factor_name", factor)
            .limit(1).execute().data)
    if rows:
        val = rows[0].get("elasticity_value")
        return float(val) if val is not None else None
    return None

def upsert_run(run_id: str, title: str, models: dict, session_id: Optional[str]):
    # session_id 是 simulation_runs 的 NOT NULL 列，必须写入
    if not session_id:
        raise HTTPException(
            status_code=400,
            detail={"stage": "upsert-run", "error": "missing session_user_id; front-end must pass session.user.id"}
        )

    exist = sb.table("simulation_runs").select("run_id").eq("run_id", run_id).execute().data
    payload = {
        "run_id": run_id,
        "title": title,
        "models": models,
        "created_at": now_iso(),
        "session_id": session_id,
    }
    if exist:
        sb.table("simulation_runs").update(payload).eq("run_id", run_id).execute()
    else:
        sb.table("simulation_runs").insert(payload).execute()


def save_artifact(run_id: str, kind: str, content: bytes, filename: str) -> str:
    # 存 Storage：bucket = 'artifacts'
    bucket = "artifacts"
    ensure_bucket(bucket)
    # ← 新增：不存在则自动创建（公有）
    path = f"{run_id}/{filename}"
    if filename.endswith(".md"):
        mime = "text/markdown; charset=utf-8"
    elif filename.endswith(".csv"):
        mime = "text/csv; charset=utf-8"
    elif filename.endswith(".html"):
        mime = "text/html; charset=utf-8"
    elif filename.endswith(".pdf"):
        mime = "application/pdf"
    else:
        mime = "text/plain; charset=utf-8"

    sb.storage.from_(bucket).upload(
        path,
        content,
        {"content-type": mime, "x-upsert": "true"}
    )
    url = sb.storage.from_(bucket).get_public_url(path)
    sb.table("simulation_artifacts").insert({
        "run_id": run_id,
        "artifact_type": kind,
        "storage_url": url,
        "size_bytes": len(content),
        "sha256": sha256_bytes(content),
        "created_at": now_iso()
    }).execute()
    return url


# =============== Storage 帮助 ===============
def ensure_bucket(bucket: str):
    """
    若 bucket 不存在则创建。
    说明：当前 storage3 版本的 create_bucket 不接受关键字参数 'public'，
    这里仅按最兼容方式创建；是否公开可在 SQL/控制台单独设置。
    """
    try:
        exists = False
        for b in sb.storage.list_buckets() or []:
            name = b.get("name") if isinstance(b, dict) else getattr(b, "name", None)
            if name == bucket:
                exists = True
                break
        if not exists:
            sb.storage.create_bucket(bucket)  # 不传 public 关键字，避免 400/TypeError
    except Exception as e:
        raise HTTPException(status_code=500, detail={"stage": "ensure-bucket", "error": str(e)})

# =============== 附件上传（最小） ===============
@app.post("/simulation_v2/upload")
async def upload_attachment(run_id: str = Form(...), file: UploadFile = File(...)):
    content = await file.read()
    digest = sha256_bytes(content)
    bucket = "user_uploads"
    ensure_bucket(bucket)


    path = f"{run_id}/{digest}_{file.filename}"
    sb.storage.from_(bucket).upload(
        path,
        content,
        {"content-type": file.content_type or "application/octet-stream", "x-upsert": "true"}
    )
    public_url = sb.storage.from_(bucket).get_public_url(path)
    sb.table("run_attachments").insert({
        "run_id": run_id, "filename": file.filename, "mime": file.content_type,
        "size_bytes": len(content), "storage_url": public_url, "sha256": digest,
        "text_summary": f"{file.filename}（自动摘要占位）", "table_preview": []
    }).execute()
    return {"ok": True, "url": public_url, "sha256": digest}



# =============== Seed：解析问题 → 候选敏感性表 + 参数建议 ===============
@app.post("/simulation_v2/seed")
def seed(req: SeedRequest):
    import traceback
    stage = "init"
    try:
        run_id = req.run_id or str(uuid4())

        



        # 1) LLM 解析公司/X/Y（传入完整指标元数据，便于依据 description/aliases 判断相关性与粒度）
        stage = "llm-parse"
        company_guess = find_company_by_name_or_best(req.question or "")

        metric_rows = sb.table("metric_alias_catalog") \
                    .select("canonical_name, description, category, aliases, is_derived") \
                    .execute().data or []
        metric_catalog = [ (r.get("canonical_name") or "").strip() for r in metric_rows ]


        parse_input = json.dumps({
            "question": req.question,
            "metric_catalog": metric_catalog,
            "metric_catalog_full": metric_rows
        }, ensure_ascii=False)
        parse_txt = call_llm(SYS_PARSE, parse_input)
        try:
            parsed = json.loads(parse_txt)
        except Exception:
            parsed = {"company": company_guess, "X": [], "Y": [], "explicit_shocks": [], "thinking": []}


        stage = "canonicalize-XY"
        company = (parsed.get("company") or company_guess)

        X_raw = alias_to_canonical(as_list(parsed.get("X")))
        X_raw = [x for x in X_raw if x in set(metric_catalog)]

        Y_from_llm = alias_to_canonical(as_list(parsed.get("Y")))
        Y_raw = [y for y in Y_from_llm if y in set(metric_catalog)]

        # 若 LLM 未返回合格的 Y：仅基于同一提示词与 metric_catalog_full 再重试一次（仍然只依赖 description/aliases）
        if not Y_raw:
            retry_input = json.dumps({
                "question": req.question,
                "metric_catalog": metric_catalog,
                "metric_catalog_full": metric_rows,
                "retry_reason": "上一次未选出合适的 Y；请仅依据 description/aliases，从 catalog_full 里挑选 2-4 个与问题最相关、粒度更细的 Y；若存在 is_derived 字段，请优先 is_derived=false 的候选"
            }, ensure_ascii=False)
            retry_txt = call_llm(SYS_PARSE, retry_input)
            try:
                retry_parsed = json.loads(retry_txt)
                Y_from_llm = alias_to_canonical(as_list(retry_parsed.get("Y")))
                Y_raw = [y for y in Y_from_llm if y in set(metric_catalog)]
            except Exception:
                Y_raw = []
        # 若仍为空：保持空，交由前端/用户选择；不再使用硬编码的 pick_default_Y_from_catalog

        # 若已有 Y_raw：基于 is_derived 进行稳定重排（False 优先，其次保留 LLM 返回顺序）
        if Y_raw:
            is_derived_map = { (r.get("canonical_name") or "").strip(): bool(r.get("is_derived")) for r in metric_rows }
            order_index = {name: idx for idx, name in enumerate(Y_from_llm)}
            Y_raw = sorted(
                Y_raw,
                key=lambda n: (is_derived_map.get(n, True), order_index.get(n, 10**6))
            )

        # 显式冲击抽取并对齐因子（现在放到 parsed 之后）
        explicit_shocks: List[Dict[str, Any]] = []
        try:
            raw_shocks = parsed.get("explicit_shocks") or []
            if isinstance(raw_shocks, list):
                for s in raw_shocks:
                    fac = (s or {}).get("factor", "")
                    fac_canon = alias_to_canonical(as_list(fac))
                    delta_val = (s or {}).get("delta", 0.0)
                    try:
                        delta = float(delta_val)
                    except Exception:
                        delta = 0.0
                    unit = (s or {}).get("unit", "percent") or "percent"
                    if fac_canon and fac_canon[0]:
                        explicit_shocks.append({"factor": fac_canon[0], "delta": delta, "unit": unit})
        except Exception:
            explicit_shocks = []


        # 2) 敏感性回填（去重 (x,y) 组合；优先DB，缺失则 LLM 推断）
        stage = "lookup-sensitivity"
        candidates = []
        seen_pairs = set()
        for x in (X_raw or []):
            for y in Y_raw:
                key = (x, y)
                if key in seen_pairs:
                    continue
                seen_pairs.add(key)

                sens = lookup_sensitivity(company, y, x)
                src = "db" if sens is not None else "llm"
                note = "命中数据库" if sens is not None else "LLM生成（基于表/附件推断）"
                if sens is None:
                    inf = llm_infer_sensitivity(company, y, x, run_id=run_id)
                    sens = inf.get("beta")
                    if inf.get("rationale"):
                        note += f"｜{inf['rationale'][:60]}"

                # —— 季节性：同样遵循“查库→LLM→用户可改” —— #
                seasonal_adjust = None
                seasonality_source = None
                seasonality_note = None
                s_q1 = s_q2 = s_q3 = s_q4 = None  # ← 新增初始化
                try:
                    row_db = (sb.table("sensitivity_analysis")
                            .select("*").eq("company_name", company)
                            .eq("canonical_metric", y).limit(1).execute().data)
                    if row_db and "seasonal_adjust" in row_db[0]:
                        rdb = row_db[0]
                        seasonal_adjust = bool(rdb.get("seasonal_adjust"))
                        seasonality_source = "db"
                        seasonality_note = "命中数据库"
                        # 带回库里的季度系数（如有）
                        s_q1, s_q2, s_q3, s_q4 = rdb.get("seasonality_q1"), rdb.get("seasonality_q2"), rdb.get("seasonality_q3"), rdb.get("seasonality_q4")
                except Exception:
                    pass
                if seasonal_adjust is None:
                    s_inf = llm_infer_seasonality(company, y, run_id=run_id)
                    seasonal_adjust = bool(s_inf.get("seasonal_adjust", False))
                    seasonality_source = "llm"
                    rz = s_inf.get("rationale") or ""
                    s_q1, s_q2, s_q3, s_q4 = s_inf.get("q1"), s_inf.get("q2"), s_inf.get("q3"), s_inf.get("q4")
                    seasonality_note = f"LLM判断季节性｜{rz[:50]}"

                # 默认值
                lag_q = 1
                unit = "percent"
                try:
                    row_db = (sb.table("sensitivity_analysis")
                            .select("*").eq("company_name", company)
                            .eq("canonical_metric", y).eq("factor_name", x)
                            .limit(1).execute().data)
                    if row_db:
                        if row_db[0].get("lag_quarters") is not None:
                            lag_q = int(row_db[0]["lag_quarters"])
                        if row_db[0].get("shock_unit"):
                            unit = str(row_db[0]["shock_unit"])
                except Exception:
                    pass

                candidates.append({
                    "company_name": company,
                    "canonical_metric": y,
                    "factor_name": x,
                    "elasticity_value": sens,
                    "lag_quarters": lag_q,
                    "shock_unit": unit,
                    "source_method": src,
                    "note": note,
                    "seasonal_adjust": seasonal_adjust,
                    "seasonality_source": seasonality_source,
                    "seasonality_note": seasonality_note,
                    "seasonality_q1": s_q1, "seasonality_q2": s_q2, "seasonality_q3": s_q3, "seasonality_q4": s_q4
                    })




        # 3) 情景建议 & 建 run（政策/走势问题→附加 Google 证据）
        stage = "call-llm-scenario"
        evidence = []
        policy_keywords = ["政策","走势","regulation","policy","关税","tariff","税率","指引","指示"]
        if any(k in req.question for k in policy_keywords) and GOOGLE_API_KEY and GOOGLE_CSE_ID:
            evidence = google_search_snippets(req.question, topk=3)

        # 单位归一：bp→percent，abs 暂保留（LLM 会忽略或描述为“无法直接换算”）
        explicit_shocks_norm = []
        for s in explicit_shocks:
            d = float(s.get("delta") or 0.0)
            u = str(s.get("unit") or "percent").lower()
            if u == "bp":
                d = d / 10000.0
                u = "percent"
            explicit_shocks_norm.append({"factor": s["factor"], "delta": d, "unit": u})

        scen_prompt = {
            "用户问题": req.question,
            "公司": company,
            "自变量X": X_raw,
            "explicit_shocks": explicit_shocks_norm,  # ← 传标准化后的值
            "外部检索": evidence
        }

        scen_llm = call_llm(SYS_SCENARIO, json.dumps(scen_prompt, ensure_ascii=False))

        # —— 规范化 LLM 返回的因子名为标准名（例如把“关税上调/关税”对齐为“平均关税税率”）——
        try:
            _sj = json.loads(scen_llm)
            _ass = _sj.get("assumptions") or []
            _norm = []
            for a in _ass:
                fac = (a or {}).get("factor", "")
                canon = alias_to_canonical(as_list(fac))
                if canon:
                    a["factor"] = canon[0]
                _norm.append(a)
            _sj["assumptions"] = _norm
            scen_llm = json.dumps(_sj, ensure_ascii=False)
        except Exception:
            pass

        arima_defaults = {"p": 1, "d": 1, "q": 1, "periods": 8}
        quantiles = [0.1, 0.5, 0.9]

        stage = "upsert-run"
        upsert_run(
                run_id,
                title=f"模拟分析2 - {company}",
                models={"arima": arima_defaults, "monte_carlo": {"quantiles": quantiles}},
                session_id=req.session_user_id
            )


        thinking = []
        try:
            t1 = (parsed.get("thinking") or [])
            if isinstance(t1, list): thinking.extend([str(x) for x in t1][:10])
        except Exception:
            pass
        try:
            sj = json.loads(scen_llm)
            t2 = (sj.get("thinking") or [])
            if isinstance(t2, list): thinking.extend([str(x) for x in t2][:10])
        except Exception:
            pass

        return {
            "run_id": run_id,
            "company": company,
            "candidates": candidates,
            "arima_defaults": arima_defaults,
            "quantiles": quantiles,
            "scenario_llm_json": scen_llm,
            "web_sources": evidence,
            "thinking": thinking
        }

    except Exception as e:
        tb = traceback.format_exc()
        # 用 HTTPException 保证依然携带 CORS 头返回
        raise HTTPException(status_code=500, detail={"stage": stage, "error": str(e), "trace": tb})

# =============== 快速查找敏感性 ===============
@app.get("/simulation_v2/lookup_sensitivity")
def api_lookup_sensitivity(company: str, metric: str, factor: str, run_id: Optional[str] = None):
    """
    查库回填：β、lag、unit，以及季节性参数（seasonal_adjust + seasonality_q1~q4）。
    逻辑：
      1) 先查 (company, metric, factor) 行；若含季节性则直接回填；
      2) 若该行无季节性，则按 (company, metric) 再查任意一行的季节性作为兜底；
      3) 若库中均无，则用 LLM 推断季节性（保持你项目现有习惯）；
      4) β 若库有用库值，否则回退 LLM 推断。
    同时兼容历史列名：若表中使用了 q1_parameter~q4_parameter，也会自动读取映射为 seasonality_q1~q4。
    """
    # ---------- 1) 主行：metric+factor ----------
    row_full = (sb.table("sensitivity_analysis").select("*")
                .eq("company_name", company)
                .eq("canonical_metric", metric)
                .eq("factor_name", factor)
                .limit(1).execute().data)

    val: Optional[float] = None
    lag_q: Optional[int] = None
    unit: Optional[str] = None
    s_flag: Optional[bool] = None
    s_src: Optional[str] = None
    s_note: Optional[str] = None
    s_q1 = s_q2 = s_q3 = s_q4 = None

    def _read_q_params(d: Dict[str, Any]):
        """兼容两套字段名：seasonality_q1..q4 与 q1_parameter..q4_parameter"""
        q1 = d.get("seasonality_q1");  q2 = d.get("seasonality_q2")
        q3 = d.get("seasonality_q3");  q4 = d.get("seasonality_q4")
        if q1 is None and ("q1_parameter" in d):
            q1 = d.get("q1_parameter"); q2 = d.get("q2_parameter")
            q3 = d.get("q3_parameter"); q4 = d.get("q4_parameter")
        return q1, q2, q3, q4

    if row_full:
        rf = row_full[0]
        val = rf.get("elasticity_value")
        lag_q = rf.get("lag_quarters")
        unit = rf.get("shock_unit")
        # 该行附带季节性（若存在）
        if "seasonal_adjust" in rf:
            s_flag = bool(rf.get("seasonal_adjust"))
            s_q1, s_q2, s_q3, s_q4 = _read_q_params(rf)
            if any(v is not None for v in (s_q1, s_q2, s_q3, s_q4)):
                s_src, s_note = "db", "命中数据库(metric+factor)"

    # ---------- 2) metric 维度兜底 ----------
    if s_flag is None or not any(v is not None for v in (s_q1, s_q2, s_q3, s_q4)):
        try:
            row_any = (sb.table("sensitivity_analysis").select("*")
                       .eq("company_name", company)
                       .eq("canonical_metric", metric)
                       .limit(1).execute().data)
            if row_any:
                ra = row_any[0]
                if s_flag is None and "seasonal_adjust" in ra:
                    s_flag = bool(ra.get("seasonal_adjust"))
                q1, q2, q3, q4 = _read_q_params(ra)
                # 仅当主行未给出时，用兜底
                if s_q1 is None and q1 is not None: s_q1 = q1
                if s_q2 is None and q2 is not None: s_q2 = q2
                if s_q3 is None and q3 is not None: s_q3 = q3
                if s_q4 is None and q4 is not None: s_q4 = q4
                if s_src is None and any(v is not None for v in (q1,q2,q3,q4)):
                    s_src, s_note = "db", "命中数据库(metric兜底)"
        except Exception:
            pass

    # ---------- 3) β/lag/unit 查库结果不足时的推断 ----------
    if val is None:
        inf = llm_infer_sensitivity(company, metric, factor, run_id=run_id)
        val = inf.get("beta")
        src_method = "llm"
        note = f"LLM生成（基于表/附件推断）｜{(inf.get('rationale') or '')[:60]}"
    else:
        src_method = "db"
        note = "命中数据库"

    if lag_q is None:
        lag_q = 1
    if not unit:
        unit = "percent"

    # ---------- 4) 季节性仍缺 → LLM 推断 ----------
    if s_flag is None or (s_flag and not any(v is not None for v in (s_q1, s_q2, s_q3, s_q4))):
        try:
            s_inf = llm_infer_seasonality(company, metric, run_id=run_id)
            if s_flag is None:
                s_flag = bool(s_inf.get("seasonal_adjust", False))
            if not any(v is not None for v in (s_q1, s_q2, s_q3, s_q4)):
                s_q1, s_q2, s_q3, s_q4 = s_inf.get("q1"), s_inf.get("q2"), s_inf.get("q3"), s_inf.get("q4")
            if s_src is None:
                s_src, s_note = "llm", f"LLM判断季节性｜{(s_inf.get('rationale') or '')[:50]}"
        except Exception:
            # 留空即可，由前端/用户二次编辑
            pass

    return {
        "company": company, "metric": metric, "factor": factor,
        "elasticity_value": (float(val) if val is not None else None),
        "lag_quarters": lag_q, "shock_unit": unit,
        "source_method": src_method, "note": note,
        # 季节性回填（带 Q1-Q4）
        "seasonal_adjust": (bool(s_flag) if s_flag is not None else None),
        "seasonality_source": s_src, "seasonality_note": s_note,
        "seasonality_q1": (None if s_q1 is None else float(s_q1)),
        "seasonality_q2": (None if s_q2 is None else float(s_q2)),
        "seasonality_q3": (None if s_q3 is None else float(s_q3)),
        "seasonality_q4": (None if s_q4 is None else float(s_q4)),
    }

# =============== Run：取数 → ARIMA → MonteCarlo → 表格/报告 ===============
@app.post("/simulation_v2/run")
def run_sim(req: RunRequest):
    run_id = req.run_id
    skip_report = bool(req.skip_report)

    # 1) 公司
    company = req.sensitivity_rows[0].company_name if req.sensitivity_rows else find_company_by_name_or_best("")

    # 2) 取 Y 的历史序列；若设置了季调则先去季调
    ys = sorted(list({row.canonical_metric for row in req.sensitivity_rows}))
    need_seasonal: Dict[str, bool] = {}
    season_coeffs: Dict[str, Dict[str, Optional[float]] ] = {}
    for y in ys:
        rows_y = [r for r in req.sensitivity_rows if r.canonical_metric == y]
        need_seasonal[y] = any(bool(r.seasonal_adjust) for r in rows_y)
        # 若任意一行提供了 q1~q4，则记录下来
        for r in rows_y:
            if any(getattr(r, f"seasonality_q{i}", None) is not None for i in (1,2,3,4)):
                season_coeffs[y] = {
                    "q1": getattr(r, "seasonality_q1", None),
                    "q2": getattr(r, "seasonality_q2", None),
                    "q3": getattr(r, "seasonality_q3", None),
                    "q4": getattr(r, "seasonality_q4", None),
                }
                break

    series_map: Dict[str, List[float]] = {}
    for y in ys:
        if need_seasonal.get(y):
            if y in season_coeffs:
                c = season_coeffs[y]
                seq = seasonal_adjust_quarterly_series_with_coeffs(
                    company, y, c["q1"], c["q2"], c["q3"], c["q4"]
                )
            else:
                seq = seasonal_adjust_quarterly_series(company, y)
        else:
            seq = load_series_from_financial_metrics(company, y, max_points=40)

        if not seq: seq = []
        if len(seq) < 8:
            seq = (seq or [100.0]) + [seq[-1] if seq else 100.0] * (8 - len(seq or []))
        series_map[y] = seq



    # 3) ARIMA 基线
    arima_cfg = req.models.arima
    periods = int(arima_cfg.get("periods", req.horizon_quarters))
    baseline_map: Dict[str, Dict[str, Any]] = {}
    for y, seq in series_map.items():
        r = arima_baseline(seq, int(arima_cfg.get("p", 1)), int(arima_cfg.get("d", 1)), int(arima_cfg.get("q", 1)), periods)
        baseline_map[y] = r  # {"pred": [...], "lower95":[...], "upper95":[...]}

    # 4) 构造情景冲击字典（factor -> {pessimistic, base, optimistic}）
    # 若前端未传，按默认：悲观-5%，平缓0%，乐观+5%
    factors = sorted(list({row.factor_name for row in req.sensitivity_rows}))
    default_delta = {"pessimistic": -0.05, "base": 0.0, "optimistic": 0.05}
    deltas: Dict[str, Dict[str, float]] = {f: dict(default_delta) for f in factors}
    if req.scenario_deltas:
        for d in req.scenario_deltas:
            deltas[d.factor] = {"pessimistic": d.pessimistic, "base": d.base, "optimistic": d.optimistic}

    # 5) 生成“情景（确定性）”路径：按各情景 ΔX，结合敏感性 β
    def path_under_scenario(base_pred: List[float], label: str) -> List[float]:
        T = len(base_pred)
        mult = np.ones(T, dtype=float)
        for row in req.sensitivity_rows:
            beta = row.elasticity_value or 0.0
            if beta == 0.0:
                continue
            lag = max(0, int(row.lag_quarters or 0))
            dx = deltas.get(row.factor_name, default_delta).get(label, 0.0)
            if row.shock_unit == "percent":
                factor = (1.0 + beta * dx)
            else:
                factor = (1.0 + beta * dx)  # 其他单位暂同处理
            vec = np.array([factor] * T, dtype=float)
            if lag > 0:
                vec[:lag] = 1.0
            mult *= vec
        return (np.array(base_pred, dtype=float) * mult).tolist()

    # 6) Monte Carlo 分位路径（与“情景”分开，命名为 MC(p10/50/90)）
    samples = int(req.models.monte_carlo.get("samples", 1000))

    # 7) 汇总为宽表（行名 → 实际(近4期) + 预测(未来 periods 期)）
    # 先确定时间坐标：以第一个指标为锚
    anchor_metric = ys[0] if ys else None
    actual_labels: List[str] = []
    forecast_labels: List[str] = []
    actual_values_map: Dict[str, List[float]] = {}
    for y in ys:
        rows_y = load_series_rows(company, y, max_points=40)
        series_map[y] = [float(r["metric_value"]) for r in rows_y] or series_map.get(y, [])
        # 近4期实际
        tail_vals = series_map[y][-4:] if len(series_map[y]) >= 4 else series_map[y]
        actual_values_map[y] = tail_vals

    if anchor_metric:
        rows_anchor = load_series_rows(company, anchor_metric, max_points=40)
        if rows_anchor:
            last_y = rows_anchor[-1]["year"]; last_q = rows_anchor[-1]["quarter"]
            # 近4期标签（不带e）
            last4 = q_back(last_y, last_q, steps=4)
            actual_labels = [f"{yy}Q{qq}" for (yy, qq) in last4]
            # 未来 periods 期标签（带 e）
            forecast_labels = q_forward(last_y, last_q, steps=periods, mark_est=True)
        else:
            # 回退：仅用 e 标签
            forecast_labels = [f"E{t}(e)" for t in range(1, periods+1)]
    else:
        forecast_labels = [f"E{t}(e)" for t in range(1, periods+1)]

    wide_table: Dict[str, List[Any]] = {}
    for y in ys:
        base_pred = baseline_map[y]["pred"]
        actual_part = actual_values_map.get(y, [])

        # 决定“加回季节性”的方式
        reseason_mode = None
        qcoef = None
        add_offs = None
        if anchor_metric:
            if need_seasonal.get(y):
                if y in season_coeffs:  # 有Q1~Q4 → 乘法加回
                    c = season_coeffs[y]
                    qcoef = {1: (c["q1"] if c["q1"] is not None else 1.0),
                            2: (c["q2"] if c["q2"] is not None else 1.0),
                            3: (c["q3"] if c["q3"] is not None else 1.0),
                            4: (c["q4"] if c["q4"] is not None else 1.0)}
                    reseason_mode = "multiplicative"
                else:                  # 自动均值法 → 加性加回
                    add_offs = compute_additive_offsets(company, y)
                    reseason_mode = "additive"

        def _rz(path: List[float]) -> List[float]:
            if not anchor_metric or reseason_mode is None:
                return path
            return reseasonalize_series(path, last_y, last_q, reseason_mode, qcoef, add_offs)

        # 基线 + 情景 + MC（全部先“去季调预测”，再“加回季节性”）
        wide_table[f"ARIMA基线｜{y}"] = [float(x) for x in actual_part] + [float(x) for x in _rz(base_pred)]
        wide_table[f"情景-悲观｜{y}"] = [float(x) for x in actual_part] + [float(x) for x in _rz(path_under_scenario(base_pred, "pessimistic"))]
        wide_table[f"情景-平缓｜{y}"] = [float(x) for x in actual_part] + [float(x) for x in _rz(path_under_scenario(base_pred, "base"))]
        wide_table[f"情景-乐观｜{y}"] = [float(x) for x in actual_part] + [float(x) for x in _rz(path_under_scenario(base_pred, "optimistic"))]

        pcts = monte_carlo_paths(base_pred, req.sensitivity_rows, samples, periods, deltas=deltas, label="base")
        # 记录 MC 未加季节性的路径，供 XLSX 侧使用
        if 'mc_map' not in locals():
            mc_map = {}
        mc_map[y] = pcts

        wide_table[f"MC(p10)｜{y}"] = [float(x) for x in actual_part] + [float(x) for x in _rz(pcts["p10"])]
        wide_table[f"MC(p50)｜{y}"] = [float(x) for x in actual_part] + [float(x) for x in _rz(pcts["p50"])]
        wide_table[f"MC(p90)｜{y}"] = [float(x) for x in actual_part] + [float(x) for x in _rz(pcts["p90"])]

    # 8) 导出 CSV（两列拆分：指标 / 场景）并追加情景参数；同时导出 XLSX（带公式与加粗分隔）
    # 构造“按指标分组”的行集合
    # 行顺序：ARIMA基线 → 情景-悲观/平缓/乐观 → MC(p10/p50/p90)
    groups = []
    for y in ys:
        base_row = ("ARIMA基线", y, wide_table[f"ARIMA基线｜{y}"])
        scen_rows = [
            ("情景-悲观", y, wide_table[f"情景-悲观｜{y}"]),
            ("情景-平缓", y, wide_table[f"情景-平缓｜{y}"]),
            ("情景-乐观", y, wide_table[f"情景-乐观｜{y}"]),
        ]
        mc_rows = [
            ("MC(p10)", y, wide_table[f"MC(p10)｜{y}"]),
            ("MC(p50)", y, wide_table[f"MC(p50)｜{y}"]),
            ("MC(p90)", y, wide_table[f"MC(p90)｜{y}"]),
        ]
        groups.append([base_row] + scen_rows + mc_rows)

    # === CSV 导出 ===
    out = io.StringIO()
    writer = csv.writer(out)
    header = ["指标", "场景"] + (actual_labels if actual_labels else []) + forecast_labels
    writer.writerow(header)

    total_cols = len(header) - 2  # 除“指标/场景”
    def _fmt(v: Any) -> str:
        try:
            if v is None:
                return ""
            vf = float(v)
            if np.isnan(vf) or np.isinf(vf):
                return ""
            return f"{vf:.6f}"
        except Exception:
            return ""

    for grp in groups:
        for scen_name, metric_name, vals in grp:
            vv = list(vals)
            if len(vv) < total_cols:
                vv = vv + [float("nan")] * (total_cols - len(vv))
            elif len(vv) > total_cols:
                vv = vv[:total_cols]
            writer.writerow([metric_name, scen_name] + [_fmt(x) for x in vv])
        # CSV 无法保留粗线，仅插入空行作为分隔视觉提示
        writer.writerow([])

    # 追加情景参数区（首列）
    writer.writerow(["情景参数（ΔX，单位：percent）"])
    for f in factors:
        dd = deltas.get(f, default_delta)
        line = f"{f}｜悲观={dd['pessimistic']:+.2%}｜平缓={dd['base']:+.2%}｜乐观={dd['optimistic']:+.2%}"
        writer.writerow([line])

    csv_bytes = out.getvalue().encode("utf-8")

    # === XLSX 导出（含公式 & 粗线分隔 & Params/Seasonality 工作表） ===
    # 为 XLSX 构造“未加季节性”的分组（实际值 + 去季调预测），使季节性在 Excel 端可编辑
    groups_xlsx = []
    for y in ys:
        actual_part = actual_values_map.get(y, [])
        base_pred = list(baseline_map[y]["pred"])
        scen_rows = [
            ("情景-悲观", y, list(path_under_scenario(base_pred, "pessimistic"))),
            ("情景-平缓", y, list(path_under_scenario(base_pred, "base"))),
            ("情景-乐观", y, list(path_under_scenario(base_pred, "optimistic"))),
        ]
        mc = (mc_map or {}).get(y, {})
        mc_rows = [
            ("MC(p10)", y, list(mc.get("p10", [None]*periods))),
            ("MC(p50)", y, list(mc.get("p50", [None]*periods))),
            ("MC(p90)", y, list(mc.get("p90", [None]*periods))),
        ]
        groups_xlsx.append([
            ("ARIMA基线", y, base_pred),
            *scen_rows,
            *mc_rows
        ])

    wb = Workbook()
    ws = wb.active
    ws.title = "Scenario"

    # 写表头
    for c, v in enumerate(header, start=1):
        ws.cell(row=1, column=c, value=v)
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16

    # ---- Params（按 metric+factor 粒度） ----
    wsP = wb.create_sheet("Params")
    wsP.append(["metric", "factor", "beta", "lag_quarters", "dx_pessimistic", "dx_base", "dx_optimistic"])
    param_rows_for_metric: Dict[str, List[int]] = {y: [] for y in ys}
    # 新增：记录每个 factor 在 Params 表里“首次出现”的行号（用于情景参数附注读取 dx_*）
    factor_first_row_map: Dict[str, int] = {}
    prow = 2
    for y in ys:
        fs_for_y = sorted({r.factor_name for r in req.sensitivity_rows if r.canonical_metric == y})
        for f in fs_for_y:
            beta = 0.0
            lagq = 0
            for r in req.sensitivity_rows:
                if r.canonical_metric == y and r.factor_name == f:
                    beta = float(r.elasticity_value or 0.0)
                    lagq = int(r.lag_quarters or 0)
                    break
            dd = deltas.get(f, default_delta)
            wsP.append([y, f, beta, lagq, dd["pessimistic"], dd["base"], dd["optimistic"]])
            # 记录当前 metric 的该行号
            param_rows_for_metric[y].append(prow)
            # 仅在“该 factor 首次出现”时登记行号（dx_* 对于不同 metric 是一致的）
            if f not in factor_first_row_map:
                factor_first_row_map[f] = prow
            prow += 1


    # ---- Seasonality（每个 metric 的 q1~q4，可在 XLSX 里编辑）----
    wsS = wb.create_sheet("Seasonality")
    wsS.append(["metric", "q1", "q2", "q3", "q4"])
    seas_row_map: Dict[str, int] = {}
    for i, y in enumerate(ys, start=2):
        if y in season_coeffs:
            c = season_coeffs[y]
            q1 = c.get("q1", 1.0) or 1.0
            q2 = c.get("q2", 1.0) or 1.0
            q3 = c.get("q3", 1.0) or 1.0
            q4 = c.get("q4", 1.0) or 1.0
        else:
            q1 = q2 = q3 = q4 = 1.0
        wsS.append([y, q1, q2, q3, q4])
        seas_row_map[y] = i
    wsS.column_dimensions['A'].width = 18

    thin = Side(style="thin", color="999999")
    thick = Side(style="thick", color="333333")

    current_row = 2
    forecast_start_col = 2 + (len(actual_labels) if actual_labels else 0) + 1  # A=1,B=2,C=3...
    # 锚定季度：若无历史，默认 4 季度作为上一期
    last_q_for_formula = 4
    try:
        if anchor_metric:
            rows_anchor = load_series_rows(company, anchor_metric, max_points=40)
            if rows_anchor:
                last_q_for_formula = int(rows_anchor[-1]["quarter"])
    except Exception:
        pass

    # 写入每个指标的组
    for grp in groups_xlsx:
        grp_start_row = current_row
        base_row_idx = None
        metric_name_for_group = grp[0][1]  # 第一行的 metric 名称
        seas_row = seas_row_map.get(metric_name_for_group)

        # 先写“实际列 + 预测列（去季调数值作为常量）”；稍后覆盖预测列为公式
        for scen_name, metric_name, pred_vals in grp:
            ws.cell(row=current_row, column=1, value=metric_name)  # 指标
            ws.cell(row=current_row, column=2, value=scen_name)    # 场景

            # 写实际列（等于 CSV 的最近 4 期实际）
            actual_part = actual_values_map.get(metric_name, [])
            col = 3
            for v in actual_part:
                ws.cell(row=current_row, column=col, value=(None if v is None else float(v)))
                col += 1

            # 写预测列为“常数占位”，随后会转为公式
            for j, v in enumerate(pred_vals, start=forecast_start_col):
                ws.cell(row=current_row, column=j, value=(None if v is None else float(v)))

            if scen_name == "ARIMA基线":
                base_row_idx = current_row
            current_row += 1

        # 覆盖“基线行”的预测列：把常数改为  常数 × 季节因子(Seasonality)
        assert base_row_idx is not None
        if seas_row:
            for col in range(forecast_start_col, 2 + len(actual_labels or []) + 1 + len(forecast_labels)):
                # 取原常数（基线去季调预测）作为字面常量
                const_val = ws.cell(row=base_row_idx, column=col).value
                if const_val is None:
                    continue
                q_expr = f"MOD({last_q_for_formula} + COLUMN() - {forecast_start_col},4)+1"
                q1 = f"Seasonality!$B${seas_row}"
                q2 = f"Seasonality!$C${seas_row}"
                q3 = f"Seasonality!$D${seas_row}"
                q4 = f"Seasonality!$E${seas_row}"
                seas = f"CHOOSE({q_expr},{q1},{q2},{q3},{q4})"
                ws.cell(row=base_row_idx, column=col, value=f"={float(const_val)}*{seas}")

        # 覆盖三类“情景-XXX”的预测区：等于“基线（已含季节性） * ∏_{(y,f)} (1+β_{y,f}·ΔX_f)”
        # 覆盖三类“情景-XXX”的预测区（仅覆盖预测列；实际列等于基线值）
        for i in range(grp_start_row+1, grp_start_row+4):  # 紧跟基线的3行是情景-悲观/平缓/乐观
            scen_name_cell = ws.cell(row=i, column=2).value or ""
            if not str(scen_name_cell).startswith("情景-"):
                continue
            # 选取 dx 列：悲观=E(5), 平缓=F(6), 乐观=G(7)
            dx_col_idx = 6
            if "悲观" in scen_name_cell: dx_col_idx = 5
            if "乐观" in scen_name_cell: dx_col_idx = 7
            dx_col_letter = get_column_letter(dx_col_idx)

            # 实际列等于基线（不带季节性处理，因为实际就是历史）
            for col in range(3, forecast_start_col):
                ws.cell(row=i, column=col, value=f"=Scenario!{get_column_letter(col)}{base_row_idx}")

            # 预测列公式（显式季节性）：
            # = [去季调基线常数]*Seasonality(q列) * Π_{(y,f)} IF(列序>lag, 1+beta*dx, 1)
            rows_for_metric = param_rows_for_metric.get(metric_name_for_group, [])
            for col in range(forecast_start_col, 2 + len(actual_labels or []) + 1 + len(forecast_labels)):
                # 该列对应的季度：以最后观测季度 last_q_for_formula 为基，将预测首列视为 next quarter
                q_expr = f"MOD({last_q_for_formula} + COLUMN() - {forecast_start_col},4)+1"
                q1 = f"Seasonality!$B${seas_row}"
                q2 = f"Seasonality!$C${seas_row}"
                q3 = f"Seasonality!$D${seas_row}"
                q4 = f"Seasonality!$E${seas_row}"
                seas = f"CHOOSE({q_expr},{q1},{q2},{q3},{q4})"

                # 基线行该列（已含季节性）的引用
                base_with_seas = f"Scenario!{get_column_letter(col)}{base_row_idx}"
                # 还原“去季调基线常数”
                base_const = f"({base_with_seas})/({seas})"

                # 情景乘子（仅当前指标的 (metric,factor) 行）
                terms = []
                for prow in rows_for_metric:
                    beta_ref = f"Params!$C${prow}"
                    lag_ref  = f"Params!$D${prow}"
                    dx_ref   = f"Params!${dx_col_letter}${prow}"
                    terms.append(
                        f"IF(COLUMN()>={forecast_start_col},IF(COLUMN()-{forecast_start_col}+1>{lag_ref},1+{beta_ref}*{dx_ref},1),1)"
                    )
                scen_multiplier = "*".join(terms) if terms else "1"

                # 最终：= base_const * seas * scen_multiplier
                formula = f"=({base_const})*({seas})*({scen_multiplier})"
                ws.cell(row=i, column=col, value=formula)


        # 组尾加粗线分隔
        last_row_in_grp = current_row - 1
        for col in range(1, 2 + (len(actual_labels) if actual_labels else 0) + 1 + len(forecast_labels)):
            cell = ws.cell(row=last_row_in_grp, column=col)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thick)


    # 情景参数附注（放在 Scenario 表最后）
    ws.cell(row=current_row+1, column=1, value="情景参数（ΔX，单位：percent）")
    rr = current_row + 2
    for f in factors:
        # 使用在 Params 构建阶段记录的“该 factor 首行”
        prow = factor_first_row_map.get(f)
        if not prow:
            continue  # 理论上不会发生；若无则跳过该 factor
        # 注意：Params 表头为 [metric, factor, beta, lag, dx_pessimistic, dx_base, dx_optimistic]
        dx_p = wsP.cell(row=prow, column=5).value
        dx_b = wsP.cell(row=prow, column=6).value
        dx_o = wsP.cell(row=prow, column=7).value
        ws.cell(
            row=rr, column=1,
            value=f"{f}｜悲观={dx_p:+.2%}｜平缓={dx_b:+.2%}｜乐观={dx_o:+.2%}"
        )
        rr += 1


    # 保存 XLSX
    xbuf = io.BytesIO()
    wb.save(xbuf)
    xlsx_bytes = xbuf.getvalue()

    # 9) 保存 run + 产物
    upsert_run(
        run_id,
        title=f"模拟分析2 - {company}",
        models=req.models.model_dump(),
        session_id=req.session_user_id
    )

    url_csv = save_artifact(run_id, "csv", csv_bytes, "scenario_table.csv")
    url_xlsx = save_artifact(run_id, "xlsx", xlsx_bytes, "scenario_table.xlsx")

    if skip_report:
        prev = sb.table("simulation_artifacts").select("*") \
            .eq("run_id", run_id).eq("artifact_type", "md") \
            .order("created_at", desc=True).limit(1).execute().data
        url_md = prev[0]["storage_url"] if prev else None
    else:
        scenario_list = [d.model_dump() for d in (req.scenario_deltas or [])]
        md_input = {
            "question": "（本次问题见 seed 时的入参）",
            "company": company,
            "models": req.models.model_dump(),
            "sensitivities": [r.model_dump() for r in req.sensitivity_rows],
            "scenario_deltas": scenario_list,
        }
        md_prompt = f"数据概览：{json.dumps(md_input, ensure_ascii=False)[:1500]} ..."
        md = call_llm(SYS_REPORT, md_prompt)
        md_bytes = md.encode("utf-8")
        url_md = save_artifact(run_id, "md", md_bytes, "report.md")
    # 生成进度型思考日志（由 LLM 自述步骤）
    progress_prompt = {
        "company": company,
        "metrics": ys,
        "actual_periods": actual_labels,
        "forecast_periods": forecast_labels,
        "scenarios": list(deltas.keys()),
        "models": req.models.model_dump()
    }
    progress_sys = """你是进度播报助手。用第一人称、口吻亲切（以“我正在…”开头），给出 3-8 条进度短句（每条<=40字）。
只返回JSON：{"thinking":["我正在…","我已完成…","我准备…"]}"""

    thinking = []
    try:
        ttxt = call_llm(progress_sys, json.dumps(progress_prompt, ensure_ascii=False))
        tjson = json.loads(ttxt)
        thinking = [str(x) for x in (tjson.get("thinking") or [])][:10]
    except Exception:
        pass

    return {
        "run_id": run_id,
        "company": company,
        "horizon": periods,
        "wide_table_url": url_csv,
        "xlsx_url": url_xlsx,          # 新增：返回 xlsx
        "report_url": url_md,
        "table_preview": header[:8],
        "sample_rows": list(wide_table.items())[:3],
        "thinking": thinking
    }


# =============== 历史 & 产物 ===============
@app.get("/simulation_v2/history")
def history():
    rows = sb.table("simulation_runs").select("*").order("created_at", desc=True).limit(50).execute().data
    return {"runs": rows}

@app.get("/simulation_v2/artifacts")
def artifacts(run_id: str):
    rows = sb.table("simulation_artifacts").select("*").eq("run_id", run_id).order("created_at", desc=True).execute().data
    return {"artifacts": rows}
